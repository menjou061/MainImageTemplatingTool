from pathlib import Path
import os

from openpyxl import load_workbook


downloads = Path(r"C:\Users\123\Downloads")
report = Path(r"C:\Users\Public\smb_interactive_probe.txt")
workbooks = [
    path
    for path in downloads.rglob("*.xlsx")
    if not path.name.startswith(("._", ".~", "~"))
]
workbook_path = max(workbooks, key=lambda path: path.stat().st_size)
workbook = load_workbook(workbook_path, read_only=True, data_only=False)

material_path = ""
for worksheet in workbook.worksheets:
    for row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row, 2).value
        if isinstance(value, str) and value.startswith("\\\\"):
            material_path = value.strip()
            break
    if material_path:
        break

report.write_text(
    "workbook=" + str(workbook_path) + "\n"
    "material=" + material_path + "\n"
    "isfile=" + str(os.path.isfile(material_path)) + "\n",
    encoding="utf-8",
)
