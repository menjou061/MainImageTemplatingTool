#!/usr/bin/env python3
"""List visible sheets or product names in an Excel workbook for the Windows L0 launcher."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = str(Path(__file__).resolve().parent)
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from channel_profile import ProfileError, get_profile


VERTICAL_FIELD_NAMES = {
    "变量名称",
    "图片目录路径",
    "图片路径",
    "图片文件路径",
    "商品图",
    "商品图片",
}

HYGIENE_RECORD_HEADERS = {"活动", "系列", "主卖点", "堆品路径", "片数", "价格条"}
HYGIENE_STANDARD_HEADERS = {"是否出图", "输出规格", "输出文件名", "产品图路径", "主卖点"}


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def as_text(value) -> str:
    return "" if value is None else str(value).strip()


def is_display_image_formula(value) -> bool:
    if not isinstance(value, str):
        return False
    normalized = "".join(value.split()).lower()
    return normalized.startswith("=") and "dispimg" in normalized


def looks_like_vertical_header(value) -> bool:
    text = as_text(value)
    return text in VERTICAL_FIELD_NAMES or text.startswith("变量")


def ensure_dimensions(worksheet) -> tuple[int, int]:
    """Populate dimensions for streamed workbooks that omit the XML range."""
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.calculate_dimension(force=True)
    return int(worksheet.max_row or 0), int(worksheet.max_column or 0)


def first_row_headers(worksheet) -> list[str]:
    _, max_column = ensure_dimensions(worksheet)
    return [as_text(worksheet.cell(1, column).value) for column in range(1, max_column + 1)]


def is_vertical_layout(worksheet) -> bool:
    """Detect the channel format with one row per product.

    The standard workbook is transposed: field names are in column A and
    product names are in row 1.  New channel sheets put field names in row 1
    and product IDs in column A.  Requiring at least two variable-like headers
    keeps ordinary product names from being mistaken for a schema.
    """
    max_row, max_column = ensure_dimensions(worksheet)
    headers = {as_text(worksheet.cell(1, column).value) for column in range(1, max_column + 1)}
    if "文件名称" in headers and any(
        header in headers for header in ("产品（精确到图片名）", "产品图路径", "图片目录路径")
    ):
        return max_row >= 2 and max_column >= 2
    row_score = sum(
        1
        for column in range(1, max_column + 1)
        if looks_like_vertical_header(worksheet.cell(1, column).value)
    )
    return row_score >= 2 and max_row >= 2 and max_column >= 2


def visible_record_products(
    worksheet,
    headers: list[str],
    profile: dict[str, Any] | None = None,
):
    max_row, _ = ensure_dimensions(worksheet)
    header_index = {name: index + 1 for index, name in enumerate(headers) if name}
    if HYGIENE_STANDARD_HEADERS.issubset(header_index):
        expected_channel = as_text((profile or {}).get("channel"))
        expected_spec = as_text((profile or {}).get("output_label"))
        for row in range(2, max_row + 1):
            if worksheet.row_dimensions[row].hidden:
                continue
            if as_text(worksheet.cell(row, header_index["是否出图"]).value) != "是":
                continue
            if expected_channel and "渠道" in header_index:
                if as_text(worksheet.cell(row, header_index["渠道"]).value) != expected_channel:
                    continue
            if expected_spec:
                row_spec = as_text(worksheet.cell(row, header_index["输出规格"]).value)
                # Product selection must use the same strict contract as
                # clean_data: blank and unknown specifications are not shown
                # for the current variant.
                if row_spec != expected_spec:
                    continue
            yield worksheet.cell(row, header_index["输出文件名"]).value
        return
    if HYGIENE_RECORD_HEADERS.issubset(header_index):
        for row in range(2, max_row + 1):
            if worksheet.row_dimensions[row].hidden:
                continue
            activity = as_text(worksheet.cell(row, header_index["活动"]).value) or "未命名活动"
            series = as_text(worksheet.cell(row, header_index["系列"]).value) or "未命名系列"
            if any(not is_blank(worksheet.cell(row, column).value) for column in range(1, len(headers) + 1)):
                yield f"{activity}_{series}_{row - 1:02d}"


def list_sheets(workbook_path: Path) -> int:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state == "visible" and worksheet.title != "WpsReserved_CellImgList":
            print(worksheet.title)
    return 0


def list_products(
    workbook_path: Path,
    sheet_name: str,
    profile_id: str | None = None,
    variant: str | None = None,
) -> int:
    # Read cached formula results where available so a product/image formula is
    # not shown to the designer as the literal formula text.
    # Normal mode exposes hidden-row metadata, which is part of the business
    # rule for product selection. It also avoids streamed sheets with no cached
    # dimension raising before the designer sees the product list.
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet 不存在：{sheet_name}", file=sys.stderr)
        return 2
    worksheet = workbook[sheet_name]
    if worksheet.sheet_state != "visible" or worksheet.title == "WpsReserved_CellImgList":
        print(f"Sheet 不可处理：{sheet_name}", file=sys.stderr)
        return 2

    max_row, max_column = ensure_dimensions(worksheet)
    headers = first_row_headers(worksheet)
    profile = get_profile(profile_id, variant) if profile_id else None
    if HYGIENE_STANDARD_HEADERS.issubset(headers) or HYGIENE_RECORD_HEADERS.issubset(headers):
        product_values = visible_record_products(worksheet, headers, profile)
    elif is_vertical_layout(worksheet):
        product_values = (
            worksheet.cell(row, 1).value
            for row in range(2, max_row + 1)
            if not worksheet.row_dimensions[row].hidden
        )
    else:
        product_values = (worksheet.cell(1, column).value for column in range(2, max_column + 1))
    seen: set[str] = set()
    for value in product_values:
        if is_blank(value) or is_display_image_formula(value):
            continue
        product = as_text(value)
        if product in seen:
            continue
        seen.add(product)
        print(product)
    return 0


def resolve_variant_for_workbook(
    workbook_path: Path,
    sheet_name: str,
    profile_id: str,
) -> str:
    """Resolve a shared record-row Sheet from its unique output spec."""
    profile = get_profile(profile_id)
    output_label_variants = {
        variant_id: as_text(config.get("output_label"))
        for variant_id, config in profile.get("variants", {}).items()
        if as_text(config.get("output_label"))
    }
    if not output_label_variants:
        return str(profile.get("default_variant") or "")

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ProfileError("E_PROFILE_SHEET_MISMATCH", f"Sheet 不存在：{sheet_name}")
    worksheet = workbook[sheet_name]
    if worksheet.sheet_state != "visible" or worksheet.title == "WpsReserved_CellImgList":
        raise ProfileError("E_PROFILE_SHEET_MISMATCH", f"Sheet 不可处理：{sheet_name}")
    headers = first_row_headers(worksheet)
    header_index = {name: index + 1 for index, name in enumerate(headers) if name}
    if not HYGIENE_STANDARD_HEADERS.issubset(header_index):
        return str(profile.get("default_variant") or "")

    expected_channel = as_text(profile.get("channel"))
    specs: set[str] = set()
    max_row, _ = ensure_dimensions(worksheet)
    for row in range(2, max_row + 1):
        if worksheet.row_dimensions[row].hidden:
            continue
        if as_text(worksheet.cell(row, header_index["是否出图"]).value) != "是":
            continue
        if expected_channel and "渠道" in header_index:
            if as_text(worksheet.cell(row, header_index["渠道"]).value) != expected_channel:
                continue
        spec = as_text(worksheet.cell(row, header_index["输出规格"]).value)
        if spec:
            specs.add(spec)

    if len(specs) != 1:
        detail = "为空" if not specs else "、".join(sorted(specs))
        raise ProfileError(
            "E_PROFILE_SHEET_MISMATCH",
            f"Sheet {sheet_name!r} 未匹配到唯一输出规格（{detail}），禁止猜测模板规格。",
        )
    spec = next(iter(specs))
    matches = [variant_id for variant_id, label in output_label_variants.items() if label == spec]
    if len(matches) != 1:
        raise ProfileError("E_PROFILE_SHEET_MISMATCH", f"输出规格 {spec!r} 未匹配到唯一模板规格。")
    return matches[0]


def main() -> int:
    if len(sys.argv) < 2:
        print("用法：l0_list_sheets.py <xlsx> 或 l0_list_sheets.py --products <xlsx> <sheet> [--profile ID --variant ID]", file=sys.stderr)
        return 2

    resolve_mode = sys.argv[1] == "--resolve-variant"
    product_mode = sys.argv[1] == "--products"
    if resolve_mode:
        if len(sys.argv) != 6 or sys.argv[4] != "--profile":
            print("用法：l0_list_sheets.py --resolve-variant <xlsx> <sheet> --profile <ID>", file=sys.stderr)
            return 2
        workbook_path = Path(sys.argv[2])
        if not workbook_path.is_file():
            print(f"Excel 文件不存在：{workbook_path}", file=sys.stderr)
            return 2
        try:
            print(resolve_variant_for_workbook(workbook_path, sys.argv[3], sys.argv[5]))
            return 0
        except ProfileError as error:
            print(str(error), file=sys.stderr)
            return 2
    if product_mode and len(sys.argv) not in {4, 6, 8}:
        print("用法：l0_list_sheets.py --products <xlsx> <sheet> [--profile ID --variant ID]", file=sys.stderr)
        return 2
    if not product_mode and len(sys.argv) != 2:
        print("用法：l0_list_sheets.py <xlsx>", file=sys.stderr)
        return 2
    workbook_path = Path(sys.argv[2] if product_mode else sys.argv[1])
    if not workbook_path.is_file():
        print(f"Excel 文件不存在：{workbook_path}", file=sys.stderr)
        return 2

    if product_mode:
        options = dict(zip(sys.argv[4::2], sys.argv[5::2]))
        unknown = set(options) - {"--profile", "--variant"}
        if unknown or ("--variant" in options and "--profile" not in options):
            print("商品列表参数不完整或不支持。", file=sys.stderr)
            return 2
        try:
            return list_products(
                workbook_path,
                sys.argv[3],
                profile_id=options.get("--profile"),
                variant=options.get("--variant"),
            )
        except ProfileError as error:
            print(str(error), file=sys.stderr)
            return 2
    return list_sheets(workbook_path)


if __name__ == "__main__":
    raise SystemExit(main())
