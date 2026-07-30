#!/usr/bin/env python3
"""Convert explicitly profiled e-commerce Excel layouts into UTF-8 CSV rows."""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# The embedded Windows Python runtime reads ``python311._pth`` and therefore
# does not automatically add this script's folder to ``sys.path``.
SCRIPT_DIR = str(Path(__file__).resolve().parent)
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from channel_profile import ProfileError, get_profile, map_vertical_headers


NAME_MAP = {
    "文件名": "商品文件名",
    "时间": "活动时间",
    "满129可用": "券门槛",
    "价格一": "价格1",
    "价格二": "价格2",
    "堆图": "商品图",
    "大尺寸": "大尺寸图",
    "DT17090-24旧": "旧包装图",
    "正式618新旧包装底": "新旧包装底图",
}
REQUIRED_IMAGE_FIELDS = {"商品图"}
REQUIRED_RECORD_FIELDS = {"活动时间", "到手", "卖点", "规格"}
STANDARD_COLUMNS = [
    "商品文件名",
    "商品图",
    "折扣",
    "券名",
    "券门槛",
    "优惠券开关",
    "活动时间",
    "到手",
    "价格1",
    "价格2",
    "卖点",
    "规格",
    "备注",
    "片数套",
    "片数数量",
    "到手标签",
    "价格活动价",
    "价格活动价副标",
    "价格优惠券",
    "价格优惠券副标",
    "价格立减",
    "价格立减副标",
    "赠品图1",
    "赠品图2",
    "赠品图3",
    "赠品文案1",
    "赠品文案2",
    "赠品文案3",
    "版式组",
    "预检异常",
    "预检提醒",
]

LEGACY_RECORD_ROW_HEADERS = [
    "活动", "系列", "主卖点", "备注", "时间", "堆品路径", "片数",
    "主推数", "价格条", "赠品路径", "赠品文案", "代言IP路径",
]
STANDARD_RECORD_ROW_HEADERS = [
    "渠道", "是否出图", "输出规格", "输出文件名", "商品SKU", "商品系列",
    "模板版式", "产品图路径", "代言人素材路径", "主卖点", "到手套数文案", "到手数量",
    "到手价", "活动价", "商品券金额", "官方立减金额", "官方立减文案", "活动时间",
    "赠品版式", "赠品文案", "赠品素材路径", "赠品说明", "设计备注", "检查状态",
]
# Kept for callers and fixtures created against the first record-row draft.
RECORD_ROW_HEADERS = LEGACY_RECORD_ROW_HEADERS


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def has_newline(value: Any) -> bool:
    return isinstance(value, str) and ("\n" in value or "\r" in value)


def is_display_image_formula(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    normalized = "".join(value.split()).lower()
    return normalized.startswith("=") and "dispimg" in normalized


def looks_like_vertical_header(value: Any) -> bool:
    text = as_text(value)
    return text in {"变量名称", "图片目录路径", "图片路径", "图片文件路径", "商品图", "商品图片"} or text.startswith("变量")


def is_vertical_layout(ws) -> bool:
    headers = {as_text(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)}
    if "文件名称" in headers and any(
        header in headers for header in ("产品（精确到图片名）", "产品图路径", "图片目录路径")
    ):
        return ws.max_row >= 2 and ws.max_column >= 2
    row_score = sum(
        1
        for column in range(1, ws.max_column + 1)
        if looks_like_vertical_header(ws.cell(1, column).value)
    )
    return row_score >= 2 and ws.max_row >= 2 and ws.max_column >= 2


def field_name(value: Any) -> str:
    raw = as_text(value)
    return NAME_MAP.get(raw, raw)


def image_basename(value: str) -> str:
    return Path(value.replace("\\", "/")).name


def normalize_image_reference(value: str) -> str:
    """Keep the Excel image path intact for Photoshop to open directly.

    The approved workbooks carry a full UNC path in each image cell.  Keeping
    that address avoids recursively scanning an entire shared drive just to
    locate one file, and makes the preflight result deterministic.
    """
    text = as_text(value)
    if os.name == "nt" or not text.startswith("\\\\"):
        return text
    # Mac regression runs can use the already-mounted SMB share. Windows keeps
    # the original UNC path so Photoshop receives the path from Excel.
    marker = "\\个护设计中心\\"
    if marker in text:
        suffix = text.split(marker, 1)[1].replace("\\", "/")
        mounted = "/Volumes/个护设计中心/" + suffix
        if os.path.isfile(mounted):
            return mounted
    return text


def material_exists(value: str) -> bool:
    text = as_text(value)
    if os.path.isfile(text):
        return True
    if text.startswith("\\\\") and "\\个护设计中心\\" in text:
        suffix = text.split("\\个护设计中心\\", 1)[1].replace("\\", "/")
        return os.path.isfile("/Volumes/个护设计中心/" + suffix)
    return False


def evaluate_concat_formula(formula: Any, ws_values, row: int, column: int) -> str:
    """Resolve the simple ``=K2&"\\"&J2`` formulas used by channel sheets.

    Excel normally stores a cached result and ``data_only=True`` supplies it.
    This fallback covers files saved without a cache, without attempting to
    become a general Excel formula engine.
    """
    if not isinstance(formula, str) or not formula.lstrip().startswith("="):
        return as_text(formula)
    cached = ws_values.cell(row, column).value if ws_values is not None else None
    if cached is not None and not (isinstance(cached, str) and cached.startswith("=")):
        return as_text(cached)
    expression = formula.strip()[1:]
    pieces: list[str] = []
    for token in expression.split("&"):
        token = token.strip()
        if len(token) >= 2 and token[0] == '"' and token[-1] == '"':
            pieces.append(token[1:-1].replace('""', '"'))
            continue
        if re.fullmatch(r"[A-Za-z]{1,3}[0-9]+", token):
            from openpyxl.utils.cell import coordinate_to_tuple

            ref_row, ref_column = coordinate_to_tuple(token.upper())
            pieces.append(as_text(ws_values.cell(ref_row, ref_column).value if ws_values is not None else ""))
            continue
        return as_text(cached if cached is not None else formula)
    return "".join(pieces)


def is_absolute_material_path(value: str) -> bool:
    text = as_text(value)
    return (
        text.startswith("\\\\")
        or text.startswith("//")
        or bool(re.match(r"^[A-Za-z]:[\\/]", text))
    )


def split_price(value: Any) -> tuple[str, str]:
    text = as_text(value).replace("￥", "").replace(" ", "")
    match = re.fullmatch(r"(\d+)(\.\d+)?", text)
    if not match:
        return text, ""
    return match.group(1), match.group(2) or ""


def price_format_error(record: dict[str, str]) -> str:
    """A price is rendered by two independent PSD text layers.

    Price2 is deliberately free-form: it may be a decimal suffix, a unit,
    discount text, or campaign-specific copy.  The data contract only requires
    both layers to receive a value.
    """
    price1 = as_text(record.get("价格1", ""))
    price2 = as_text(record.get("价格2", ""))
    if not price1 or not price2:
        return "价格1、价格2必须完整填写"
    return ""


def is_disabled_image(value: Any) -> bool:
    text = as_text(value)
    return is_blank(text) or image_basename(text).lower() in {"无", "无.png", "none", "null"}


def normalize_price_part(value: Any, part: str) -> str:
    text = as_text(value)
    if part == "价格2" and re.fullmatch(r"0\.\d+", text):
        return text[1:]
    return text


def validate_variable_name(name: str) -> None:
    if name and (name[0].isdigit() or re.search(r"[\\/:,;\r\n]", name)):
        raise ValueError(f"变量名不符合 Photoshop/CSV 规范：{name}")


def choose_inputs(args: argparse.Namespace) -> tuple[Path, str, Path]:
    workbook_path = Path(args.xlsx).expanduser() if args.xlsx else None
    if workbook_path is None:
        try:
            from tkinter import Tk, filedialog

            root = Tk()
            root.withdraw()
            picked = filedialog.askopenfilename(title="选择 Excel 变量表", filetypes=[("Excel", "*.xlsx *.xltx")])
            root.destroy()
            workbook_path = Path(picked) if picked else None
        except Exception:
            workbook_path = Path(input("请输入 Excel 文件路径：").strip())
    if not workbook_path or not workbook_path.is_file():
        raise SystemExit("未找到 Excel 文件。")

    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        raise SystemExit("工作簿中没有可用的可见 Sheet。")
    sheet = args.sheet
    if not sheet:
        print("可用 Sheet：")
        for index, title in enumerate(visible_sheets, 1):
            print(f"  {index}. {title}")
        selection = input("请输入 Sheet 编号或名称：").strip()
        sheet = visible_sheets[int(selection) - 1] if selection.isdigit() else selection
    if sheet not in visible_sheets:
        raise SystemExit("所选 Sheet 不存在、已隐藏，或是 WPS 内嵌图片索引 Sheet。")

    output_dir = Path(args.output_dir).expanduser() if args.output_dir else workbook_path.parent / f"套版数据_{sheet}"
    return workbook_path, sheet, output_dir


def source_variables(ws, layout: str = "horizontal", profile: dict[str, Any] | None = None) -> list[tuple[int, str]]:
    variables: list[tuple[int, str]] = []
    if layout == "vertical":
        headers = [as_text(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
        header_mapping = map_vertical_headers(headers, profile or {})
        for column in range(1, ws.max_column + 1):
            raw = as_text(ws.cell(1, column).value)
            if raw not in header_mapping:
                continue
            name = header_mapping[raw]
            validate_variable_name(name)
            variables.append((column, name))
        return variables
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row, 1).value
        if is_blank(raw) or is_display_image_formula(raw):
            continue
        raw_name = as_text(raw)
        # These are worksheet notes, not data fields for a product row.
        if raw_name.startswith("\\\\") or raw_name.startswith("//"):
            continue
        name = (profile or {}).get("mapping", {}).get(raw_name, field_name(raw_name))
        validate_variable_name(name)
        variables.append((row, name))
    return variables


def row_from_column(ws, column: int, variables: list[tuple[int, str]], ws_values=None) -> dict[str, str]:
    row: dict[str, str] = {}
    for source_row, target_name in variables:
        value = ws.cell(source_row, column).value
        if isinstance(value, str) and value.lstrip().startswith("="):
            value = evaluate_concat_formula(value, ws_values, source_row, column)
        # Optional PSD variables may legitimately contain line breaks. The
        # fixed main-image contract remains strict.
        if has_newline(value) and (target_name in REQUIRED_RECORD_FIELDS or target_name in REQUIRED_IMAGE_FIELDS):
            row["预检异常"] = append_issue(row.get("预检异常", ""), "脏数据")
        text = normalize_price_part(value, target_name)
        row[target_name] = normalize_image_reference(text)

    if "价格" in row and is_blank(row.get("价格1")) and is_blank(row.get("价格2")):
        row["价格1"], row["价格2"] = split_price(row["价格"])
    elif is_blank(row.get("价格2")) and re.fullmatch(r"\d+\.\d+", row.get("价格1", "")):
        row["价格1"], row["价格2"] = split_price(row["价格1"])

    coupon_fields = [row.get("折扣", ""), row.get("券名", ""), row.get("券门槛", "")]
    filled = [not is_blank(value) for value in coupon_fields]
    if not any(filled):
        row["优惠券开关"] = "否"
    else:
        row["优惠券开关"] = "是"
        if not all(filled):
            row["预检异常"] = append_issue(row.get("预检异常", ""), "字段为空")

    return row


def row_from_vertical(
    ws,
    ws_values,
    source_row: int,
    variables: list[tuple[int, str]],
) -> dict[str, str]:
    row: dict[str, str] = {}
    for source_column, target_name in variables:
        formula_value = ws.cell(source_row, source_column).value
        value = formula_value
        if isinstance(formula_value, str) and formula_value.lstrip().startswith("="):
            value = evaluate_concat_formula(formula_value, ws_values, source_row, source_column)
        if has_newline(value) and (target_name in REQUIRED_RECORD_FIELDS or target_name in REQUIRED_IMAGE_FIELDS):
            row["预检异常"] = append_issue(row.get("预检异常", ""), "脏数据")
        text = normalize_price_part(value, target_name)
        row[target_name] = normalize_image_reference(text)

    if "价格" in row and is_blank(row.get("价格1")) and is_blank(row.get("价格2")):
        row["价格1"], row["价格2"] = split_price(row["价格"])
    elif is_blank(row.get("价格2")) and re.fullmatch(r"\d+\.\d+", row.get("价格1", "")):
        row["价格1"], row["价格2"] = split_price(row["价格1"])

    row["优惠券开关"] = "否"
    return row


def split_price_bar(value: Any) -> dict[str, str]:
    """Split hygiene price copy while allowing optional coupon/reduction terms."""
    text = as_text(value).replace("￥", "")
    equation = re.match(
        r"^\s*(?P<to_hand>[\d.]+)\s*[（(](?P<to_hand_label>[^）)]*)[）)]\s*=\s*(?P<terms>.+?)\s*$",
        text,
    )
    if not equation:
        return {"价格条": text}
    terms = re.findall(r"(?:^|-)\s*([\d.]+)\s*[（(]([^）)]*)[）)]", equation.group("terms"))
    if not terms or "活动价" not in terms[0][1]:
        return {"价格条": text}
    result = {
        "价格条": text,
        "到手": "¥" + equation.group("to_hand"),
        "到手标签": equation.group("to_hand_label"),
        "价格活动价": terms[0][0],
        "价格活动价副标": terms[0][1],
        "价格优惠券": "",
        "价格优惠券副标": "",
        "价格立减": "",
        "价格立减副标": "",
    }
    for amount, label in terms[1:]:
        if "券" in label and not result["价格优惠券"]:
            result["价格优惠券"] = amount
            result["价格优惠券副标"] = label
        elif ("立减" in label or "直降" in label) and not result["价格立减"]:
            result["价格立减"] = amount
            result["价格立减副标"] = label
        else:
            return {"价格条": text}
    return result


def split_piece_count(value: Any) -> dict[str, str]:
    text = as_text(value)
    match = re.search(
        r"(?:(?P<sets>\d+套)(?:含赠)?|(?P<label>单套|含赠))到手\s*(?P<count>[\d.]+(?:片|条|卷|包|抽|提|节))",
        text,
    )
    if match:
        prefix = match.group("sets") or match.group("label")
        return {"片数套": prefix + "到手", "片数数量": match.group("count")}
    return {"片数数量": text}


def split_lines(value: Any, limit: int | None = None) -> list[str]:
    lines = [line.strip() for line in as_text(value).replace("\r", "").split("\n") if line.strip()]
    return lines if limit is None else lines[:limit]


def price_text(value: Any, *, currency: bool = False) -> str:
    text = as_text(value).replace("￥", "").replace("¥", "")
    return ("¥" + text) if currency and text else text


def row_from_standard_record(
    values: dict[str, str], variant: str, output_label: str | None = None
) -> dict[str, str]:
    gift_paths = split_lines(values.get("赠品素材路径", ""))
    gift_copy = split_lines(values.get("赠品文案", ""))
    record: dict[str, str] = {
        "商品文件名": values.get("输出文件名", ""),
        "系列": values.get("商品系列", ""),
        "卖点": values.get("主卖点", ""),
        "备注": values.get("设计备注", ""),
        "活动时间": values.get("活动时间", ""),
        "商品图": normalize_image_reference(values.get("产品图路径", "")),
        "代言IP路径": normalize_image_reference(values.get("代言人素材路径", "")),
        "片数套": values.get("到手套数文案", ""),
        "片数数量": values.get("到手数量", ""),
        "到手": price_text(values.get("到手价", ""), currency=True),
        "到手标签": values.get("到手套数文案", ""),
        "价格活动价": price_text(values.get("活动价", "")),
        "价格活动价副标": "活动价",
        "价格优惠券": price_text(values.get("商品券金额", "")),
        "价格优惠券副标": "入会领商品券" if values.get("商品券金额", "") else "",
        "价格立减": price_text(values.get("官方立减金额", "")),
        "价格立减副标": values.get("官方立减文案", ""),
        "版式组": values.get("模板版式", ""),
        "赠品说明": values.get("赠品说明", ""),
        "预检异常": "",
    }
    record["代言IP"] = record["代言IP路径"]
    record["价格优惠券开关"] = "是" if record["价格优惠券"] else "否"
    record["价格立减开关"] = "是" if record["价格立减"] else "否"
    expected_spec = as_text(output_label) or as_text(variant)
    actual_spec = as_text(values.get("输出规格", ""))
    if not actual_spec:
        record["预检异常"] = append_issue(
            record["预检异常"],
            f"E_OUTPUT_SPEC_MISSING:输出规格为空，应填写{expected_spec or '当前规格'}",
        )
    elif actual_spec != expected_spec:
        record["预检异常"] = append_issue(
            record["预检异常"],
            f"E_OUTPUT_SPEC_MISMATCH:输出规格“{actual_spec}”与当前规格“{expected_spec}”不一致",
        )
    if values.get("是否出图", "") != "是":
        record["_跳过"] = "未选择出图"
    check_status = as_text(values.get("检查状态", ""))
    if check_status != "可出图":
        record["预检异常"] = append_issue(
            record["预检异常"],
            f"E_CHECK_STATUS_INVALID:检查状态“{check_status or '空'}”不是“可出图”",
        )
    gift_layout = as_text(values.get("赠品版式", ""))
    if gift_layout:
        record["预检异常"] = append_issue(
            record["预检异常"],
            f"E_GIFT_LAYOUT_UNSUPPORTED:暂不支持自动处理赠品版式“{gift_layout}”，请留空",
        )
    if len(gift_paths) != len(gift_copy) or len(gift_paths) > 3:
        record["预检异常"] = append_issue(
            record["预检异常"],
            "E_GIFT_PAIR_MISMATCH:赠品素材与文案数量不一致或超过3项",
        )
    for index in range(3):
        record[f"赠品图{index + 1}"] = normalize_image_reference(gift_paths[index]) if index < len(gift_paths) else ""
        record[f"赠品文案{index + 1}"] = gift_copy[index] if index < len(gift_copy) else ""
    return record


def row_from_record(ws, row_number: int, headers: list[str]) -> dict[str, str]:
    values = {headers[index]: as_text(ws.cell(row_number, index + 1).value) for index in range(len(headers))}
    record: dict[str, str] = {
        "活动": values.get("活动", ""),
        "系列": values.get("系列", ""),
        "卖点": values.get("主卖点", "").strip(),
        "备注": values.get("备注", "").strip(),
        "活动时间": values.get("时间", ""),
        "商品图": normalize_image_reference(values.get("堆品路径", "")),
        "主推数": values.get("主推数", ""),
        "代言IP路径": normalize_image_reference(values.get("代言IP路径", "")),
    }
    record["代言IP"] = record["代言IP路径"]
    record.update(split_piece_count(values.get("片数", "")))
    record.update(split_price_bar(values.get("价格条", "")))
    gift_paths = split_lines(values.get("赠品路径", ""))
    gift_copy = split_lines(values.get("赠品文案", ""))
    if len(gift_paths) != len(gift_copy) or len(gift_paths) > 3:
        record["预检异常"] = "E_GIFT_PAIR_MISMATCH:赠品素材与文案数量不一致或超过3项"
    for index in range(3):
        record[f"赠品图{index + 1}"] = normalize_image_reference(gift_paths[index]) if index < len(gift_paths) else ""
        record[f"赠品文案{index + 1}"] = gift_copy[index] if index < len(gift_copy) else ""
    # Activity and series are intentionally metadata only. The layout selector
    # follows the approved path rule and is recorded for the PSD task report.
    record["版式组"] = "小马无侧边" if "(IP)" in record.get("商品图", "") or "（IP）" in record.get("商品图", "") else "无代言人"
    record["价格优惠券开关"] = "是" if record.get("价格优惠券") else "否"
    record["价格立减开关"] = "是" if record.get("价格立减") else "否"
    return record


def append_issue(existing: str, issue: str) -> str:
    return issue if not existing else existing + ";" + issue


def quality_score(record: dict[str, str]) -> int:
    return sum(1 for key, value in record.items() if key not in {"预检异常", "优惠券开关"} and not is_blank(value))


def add_material_precheck(record: dict[str, str], image_fields: set[str] | None = None) -> None:
    """Validate each Excel-provided image address without scanning a folder."""
    for field in image_fields or REQUIRED_IMAGE_FIELDS:
        image_path = as_text(record.get(field, ""))
        if is_disabled_image(image_path):
            if field == "商品图":
                record["预检异常"] = append_issue(record.get("预检异常", ""), "素材地址缺失:商品图")
            continue
        if not is_absolute_material_path(image_path):
            record["预检异常"] = append_issue(record.get("预检异常", ""), f"素材地址缺失:{field}")
            continue
        if not material_exists(image_path):
            record["预检异常"] = append_issue(record.get("预检异常", ""), f"缺图:{field}")


def add_required_field_precheck(record: dict[str, str], required_fields: set[str] | None = None) -> None:
    required = required_fields if required_fields is not None else REQUIRED_RECORD_FIELDS
    missing = [field for field in required if is_blank(record.get(field, ""))]
    if missing:
        record["预检异常"] = append_issue(record.get("预检异常", ""), "字段为空:" + "、".join(sorted(missing)))


def add_filename_precheck(record: dict[str, str], image_fields: set[str] | None = None) -> None:
    for field in image_fields or REQUIRED_IMAGE_FIELDS:
        image_name = image_basename(record.get(field, ""))
        if is_disabled_image(image_name):
            continue
        if "m²" in image_name or "²" in image_name:
            record["预检异常"] = append_issue(record.get("预检异常", ""), f"文件名特殊字符:{field}")


def exception_record(record: dict[str, str], product: str, column: int, issue: str, detail: str) -> dict[str, str]:
    result = dict(record)
    code_match = re.search(r"\b(E_[A-Z_]+)\b", detail)
    if code_match:
        code = code_match.group(1)
    elif "缺图" in detail or "素材地址缺失" in detail:
        code = "E_MATERIAL_MISSING"
    elif "字段为空" in detail:
        code = "E_REQUIRED_FIELD_MISSING"
    else:
        code = "E_ROW_INVALID"
    suggestions = {
        "E_GIFT_PAIR_MISMATCH": "赠品文案和素材路径按顺序一一对应，最多3项。",
        "E_MATERIAL_MISSING": "补齐 Windows 可直接打开的完整素材路径。",
        "E_REQUIRED_FIELD_MISSING": "补齐该行必填内容后只重跑失败商品。",
        "E_RECORD_KEY_MISSING": "填写唯一的输出文件名。",
        "E_RECORD_KEY_DUPLICATE": "每行使用不同的输出文件名，避免成品互相覆盖。",
        "E_CHANNEL_MISMATCH": "把该行“渠道”改为当前所选渠道，或返回工具选择匹配的品类和渠道。",
    }
    result.update({
        "商品文件名": product,
        "源列": str(column),
        "异常类型": issue,
        "错误码": code,
        "异常详情": detail,
        "建议动作": suggestions.get(code, "按异常详情修正该行后重跑。"),
    })
    return result


def write_csv(path: Path, fieldnames: list[str], records: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def build_record_rows(
    ws,
    output_dir: Path,
    product_filter: set[str] | None,
    limit: int | None,
    profile: dict[str, Any],
) -> tuple[int, int, Path, Path, Path]:
    header_row = int(profile.get("sheet", {}).get("header_row", 1))
    headers = [as_text(ws.cell(header_row, column).value) for column in range(1, ws.max_column + 1)]
    header_set = set(headers)
    if set(STANDARD_RECORD_ROW_HEADERS).issubset(header_set):
        schema = "standard"
        required_headers = set(STANDARD_RECORD_ROW_HEADERS)
    elif set(LEGACY_RECORD_ROW_HEADERS).issubset(header_set):
        schema = "legacy"
        required_headers = set(LEGACY_RECORD_ROW_HEADERS)
    else:
        configured_sets = profile.get("sheet", {}).get("required_header_sets", [])
        candidates = [set(items) for items in configured_sets] or [set(LEGACY_RECORD_ROW_HEADERS), set(STANDARD_RECORD_ROW_HEADERS)]
        closest = min(candidates, key=lambda candidate: len(candidate - header_set))
        missing = sorted(closest - header_set)
        raise ProfileError("E_PROFILE_SCHEMA_MISMATCH", "缺少列：" + "、".join(missing))
    missing = sorted(required_headers - header_set)
    if missing:
        raise ProfileError("E_PROFILE_SCHEMA_MISMATCH", "缺少列：" + "、".join(missing))
    required_fields = set(profile.get("record_required_fields", [
        "活动时间", "卖点", "片数套", "片数数量", "到手", "价格活动价",
    ]))
    records: list[dict[str, str]] = []
    record_source_rows: list[int] = []
    seen_products: dict[str, int] = {}
    for row_number in range(header_row + 1, ws.max_row + 1):
        if ws.row_dimensions[row_number].hidden:
            continue
        if all(is_blank(ws.cell(row_number, column).value) for column in range(1, len(headers) + 1)):
            continue
        if schema == "standard":
            values = {headers[index]: as_text(ws.cell(row_number, index + 1).value) for index in range(len(headers))}
            record = row_from_standard_record(
                values,
                str(profile["variant"]),
                str(profile.get("output_label", "")),
            )
            if record.pop("_跳过", ""):
                continue
            expected_channel = as_text(profile.get("channel", ""))
            actual_channel = values.get("渠道", "")
            if actual_channel != expected_channel:
                channel_issue = (
                    f"E_CHANNEL_MISMATCH:表格渠道“{actual_channel or '空'}”"
                    f"与已选渠道“{expected_channel or '未配置'}”不一致"
                )
                existing_issue = record.get("预检异常", "")
                record["预检异常"] = append_issue(channel_issue, existing_issue) if existing_issue else channel_issue
            product = record.get("商品文件名", "")
        else:
            record = row_from_record(ws, row_number, headers)
            activity = record.get("活动", "") or "未命名活动"
            series = record.get("系列", "") or "未命名系列"
            product = f"{activity}_{series}_{row_number - header_row:02d}"
        record["商品文件名"] = product
        record["profile_id"] = profile["profile_id"]
        record["profile_version"] = profile["profile_version"]
        record["variant"] = profile["variant"]
        record.setdefault("预检异常", "")
        if product_filter and product not in product_filter:
            continue
        if is_blank(product):
            record["预检异常"] = append_issue(record.get("预检异常", ""), "E_RECORD_KEY_MISSING:输出文件名为空")
        elif product in seen_products:
            record["预检异常"] = append_issue(
                record.get("预检异常", ""),
                f"E_RECORD_KEY_DUPLICATE:输出文件名与第{seen_products[product]}行重复",
            )
        else:
            seen_products[product] = row_number
        if is_blank(record.get("商品图")):
            record["预检异常"] = append_issue(record.get("预检异常", ""), "素材地址缺失:商品图")
        elif not material_exists(record["商品图"]):
            record["预检异常"] = append_issue(record.get("预检异常", ""), "缺图:商品图")
        for index in range(1, 4):
            gift_path = record.get(f"赠品图{index}", "")
            if gift_path and not material_exists(gift_path):
                record["预检异常"] = append_issue(record.get("预检异常", ""), f"缺图:赠品图{index}")
        missing_fields = [field for field in sorted(required_fields) if is_blank(record.get(field, ""))]
        if missing_fields:
            record["预检异常"] = append_issue(record.get("预检异常", ""), "字段为空:" + "、".join(missing_fields))
        if record.get("备注") and not record["备注"].startswith("*"):
            record["预检提醒"] = "备注建议以*开头"
        records.append(record)
        record_source_rows.append(row_number)
    if limit is not None:
        if limit < 1:
            raise ValueError("--limit 必须是大于 0 的整数。")
        records = records[:limit]
        record_source_rows = record_source_rows[:limit]
    exceptions = [
        exception_record(
            record,
            record.get("商品文件名", ""),
            source_row,
            "数据预检不通过",
            record["预检异常"],
        )
        for record, source_row in zip(records, record_source_rows)
        if record.get("预检异常")
    ]
    clean_records = [record for record in records if not record.get("预检异常")]
    all_records = records
    seen_columns = set(STANDARD_COLUMNS)
    discovered_columns: list[str] = []
    for record in all_records + exceptions:
        for key in record:
            if key not in seen_columns and key not in {"源列", "异常类型", "错误码", "异常详情", "建议动作"}:
                seen_columns.add(key)
                discovered_columns.append(key)
    data_columns = STANDARD_COLUMNS + discovered_columns
    for record in all_records:
        for column in data_columns:
            record.setdefault(column, "")
    output_dir.mkdir(parents=True, exist_ok=True)
    data_path = output_dir / "data.csv"
    all_data_path = output_dir / "data_全部记录.csv"
    error_path = output_dir / "异常记录.csv"
    write_csv(data_path, data_columns, clean_records)
    write_csv(all_data_path, data_columns, all_records)
    error_columns = ["商品文件名", "源列", "异常类型", "错误码", "异常详情", "建议动作"] + [column for column in data_columns if column != "商品文件名"]
    write_csv(error_path, error_columns, exceptions)
    return len(clean_records), len(exceptions), data_path, all_data_path, error_path


def build_data(
    workbook_path: Path,
    sheet_name: str,
    output_dir: Path,
    product_filter: set[str] | None,
    limit: int | None = None,
    profile_id: str | None = None,
    variant: str | None = None,
) -> tuple[int, int, Path, Path, Path]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    values_workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    ws = workbook[sheet_name]
    ws_values = values_workbook[sheet_name]
    if ws.sheet_state != "visible" or ws.title == "WpsReserved_CellImgList":
        raise ValueError("不能处理隐藏 Sheet 或 WPS 内嵌图片索引 Sheet。")
    profile = get_profile(profile_id, variant)
    layout = "vertical" if is_vertical_layout(ws) else "horizontal"
    if profile.get("layout") == "record_rows":
        return build_record_rows(ws, output_dir, product_filter, limit, profile)
    if layout != profile["layout"]:
        raise ProfileError("E_PROFILE_SCHEMA_MISMATCH", f"profile {profile['profile_id']} 不支持此 Sheet 布局")
    variables = source_variables(ws, layout, profile)
    if profile.get("execution_mode") == "photoshop_variables":
        required_fields = {
            variable["name"]
            for variable in profile.get("required_psd_variables", [])
            if variable.get("type") == "text"
        }
        required_image_fields = {
            variable["name"]
            for variable in profile.get("required_psd_variables", [])
            if variable.get("type") == "pixel_replacement"
        }
    else:
        if layout == "horizontal":
            required_fields = REQUIRED_RECORD_FIELDS
        else:
            required_fields = {
                variable["name"]
                for variable in profile.get("required_psd_variables", [])
                if variable.get("type") == "text"
            } or {"预估到手价", "卖点", "规格"}
        required_image_fields = {
            variable["name"]
            for variable in profile.get("required_psd_variables", [])
            if variable.get("type") == "smart_object"
        } or REQUIRED_IMAGE_FIELDS
    # Keep every selected product candidate, including records with preflight
    # issues.  The UI can later let the designer choose either the clean set or
    # the original-content set without reconstructing rows from an exception
    # report (which would lose fields and their original values).
    candidates: list[tuple[int, str, dict[str, str]]] = []
    exceptions: list[dict[str, str]] = []
    if layout == "vertical":
        sku_header = profile.get("sheet", {}).get("sku_header")
        sku_column = 1
        if sku_header:
            for column in range(1, ws.max_column + 1):
                if as_text(ws.cell(1, column).value) == sku_header:
                    sku_column = column
                    break
            else:
                raise ProfileError("E_PROFILE_SCHEMA_MISMATCH", f"缺少 SKU 列：{sku_header}")
        source_items = ((row, as_text(ws_values.cell(row, sku_column).value)) for row in range(2, ws.max_row + 1))
    else:
        source_items = ((column, as_text(ws_values.cell(1, column).value)) for column in range(2, ws.max_column + 1))
    for source_index, product in source_items:
        raw_product = ws.cell(source_index, sku_column).value if layout == "vertical" else ws.cell(1, source_index).value
        if is_blank(product) or is_display_image_formula(raw_product):
            continue
        if product_filter and product not in product_filter:
            continue
        if layout == "vertical":
            record = row_from_vertical(ws, ws_values, source_index, variables)
        else:
            record = row_from_column(ws, source_index, variables, ws_values)
        record["商品文件名"] = product
        record["profile_id"] = profile["profile_id"]
        record["profile_version"] = profile["profile_version"]
        record["variant"] = profile["variant"]
        if has_newline(raw_product):
            record["预检异常"] = append_issue(record.get("预检异常", ""), "脏数据")
            exceptions.append(exception_record(record, product, source_index, "脏数据", "商品文件名包含换行符"))
        add_filename_precheck(record, required_image_fields)
        add_material_precheck(record, required_image_fields)
        add_required_field_precheck(record, required_fields)
        if "脏数据" in record.get("预检异常", "") and not has_newline(raw_product):
            exceptions.append(exception_record(record, product, source_index, "脏数据", "任一单元格包含换行符"))
        price_error = price_format_error(record)
        if price_error:
            record["预检异常"] = append_issue(record.get("预检异常", ""), "价格格式异常")
            exceptions.append(exception_record(record, product, source_index, "价格格式异常", price_error))
        precheck = record.get("预检异常", "")
        if precheck and not ("脏数据" in precheck or price_error):
            issue_type = "数据预检不通过"
            if "缺图:" in precheck:
                issue_type = "缺图"
            elif "素材地址缺失:" in precheck:
                issue_type = "素材地址缺失"
            elif "字段为空" in precheck:
                issue_type = "字段为空"
            elif "文件名特殊字符:" in precheck:
                issue_type = "文件名特殊字符"
            exceptions.append(exception_record(record, product, source_index, issue_type, precheck))
        candidates.append((source_index, product, record))

    by_product: dict[str, list[tuple[int, str, dict[str, str]]]] = defaultdict(list)
    for candidate in candidates:
        by_product[candidate[1]].append(candidate)

    clean_records: list[dict[str, str]] = []
    all_records: list[dict[str, str]] = []
    clean_record_sources: dict[str, int] = {}
    for product, group in by_product.items():
        group.sort(key=lambda item: (-quality_score(item[2]), item[0]))
        kept = group[0]
        kept_record = dict(kept[2])
        kept_record["商品文件名"] = product
        all_records.append(kept_record)
        if not kept_record.get("预检异常", ""):
            clean_records.append(kept_record)
        clean_record_sources[product] = kept[0]
        precheck = kept_record.get("预检异常", "")
        for duplicate in group[1:]:
            exceptions.append(
                exception_record(
                    duplicate[2],
                    product,
                    duplicate[0],
                    "重复商品名",
                    f"保留了信息更完整的第 {kept[0]} 列；当前为第 {duplicate[0]} 列。",
                )
            )

    if limit is not None:
        if limit < 1:
            raise ValueError("--limit 必须是大于 0 的整数。")
        clean_records = clean_records[:limit]
        selected_products = {record["商品文件名"] for record in clean_records}
        if selected_products:
            maximum_source_column = max(clean_record_sources[product] for product in selected_products)
            all_records = [record for record in all_records if record.get("商品文件名") in selected_products]
            exceptions = [
                record for record in exceptions
                if int(record.get("源列", 0) or 0) <= maximum_source_column
            ]
        else:
            all_records = []
    seen_columns = set(STANDARD_COLUMNS)
    discovered_columns = []
    for record in all_records + exceptions:
        for key in record:
            if key not in seen_columns and key not in {"源列", "异常类型", "错误码", "异常详情", "建议动作"}:
                seen_columns.add(key)
                discovered_columns.append(key)
    data_columns = STANDARD_COLUMNS + discovered_columns
    for record in clean_records + all_records:
        for column in data_columns:
            record.setdefault(column, "")

    output_dir.mkdir(parents=True, exist_ok=True)
    data_path = output_dir / "data.csv"
    all_data_path = output_dir / "data_全部记录.csv"
    error_path = output_dir / "异常记录.csv"
    write_csv(data_path, data_columns, clean_records)
    write_csv(all_data_path, data_columns, all_records)
    error_columns = ["商品文件名", "源列", "异常类型", "错误码", "异常详情", "建议动作"] + [column for column in data_columns if column != "商品文件名"]
    write_csv(error_path, error_columns, exceptions)
    return len(clean_records), len(exceptions), data_path, all_data_path, error_path


def main() -> None:
    parser = argparse.ArgumentParser(description="将横向商品变量表清洗为套版 data.csv（UTF-8）。")
    parser.add_argument("--xlsx", help="Excel 变量表路径；省略时弹窗选择。")
    parser.add_argument("--sheet", help="要处理的可见 Sheet 名称；省略时交互选择。")
    parser.add_argument("--output-dir", help="data.csv、data_全部记录.csv 与异常记录.csv 输出文件夹。")
    parser.add_argument(
        "--product",
        action="append",
        help="仅导出指定商品文件名；可重复传入以选择多个商品。",
    )
    parser.add_argument("--limit", type=int, help="仅导出前 N 个去重后的商品，用于分批试跑。")
    parser.add_argument("--profile", help="明确指定 channel profile；省略时使用 legacy-v1。")
    parser.add_argument("--variant", help="明确指定 profile variant。")
    args = parser.parse_args()
    workbook_path, sheet_name, output_dir = choose_inputs(args)
    try:
        count, errors, data_path, all_data_path, error_path = build_data(
            workbook_path,
            sheet_name,
            output_dir,
            set(args.product) if args.product else None,
            args.limit,
            args.profile,
            args.variant,
        )
    except (OSError, ValueError) as error:
        raise SystemExit(f"处理失败：{error}")
    print(f"完成：有效记录 {count} 条，异常 {errors} 条。")
    print(f"数据文件：{data_path}")
    print(f"继续生成数据文件：{all_data_path}")
    print(f"异常记录：{error_path}")


if __name__ == "__main__":
    main()
