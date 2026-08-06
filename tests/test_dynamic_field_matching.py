"""Regression coverage for v1.3 field aliases and additive fields."""
from __future__ import annotations

import csv
import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
INTERNAL = ROOT / "L0_Windows命令行版" / "_internal"
SPEC = importlib.util.spec_from_file_location("l0_clean_data_dynamic", INTERNAL / "clean_data.py")
assert SPEC and SPEC.loader
clean_data = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(clean_data)

from channel_profile import map_vertical_headers, resolve_field_name, get_profile  # noqa: E402


class DynamicFieldMatchingTest(unittest.TestCase):
    def test_price3_alias_resolves_without_script_rule(self) -> None:
        profile = get_profile("tmall-positional-v1", "main-750")
        mapping = map_vertical_headers(
            [
                "商品文件名",
                "利益点1",
                "利益点2",
                "预估到手价",
                "价格1",
                "价格2",
                "价格3",
                "规格",
                "卖点",
                "产品",
            ],
            profile,
        )
        self.assertEqual(mapping["商品文件名"], "商品文件名")
        self.assertEqual(mapping["价格3"], "价格3")
        self.assertEqual(resolve_field_name("@价格3", profile)[0], "价格3")

    def test_unknown_additive_field_is_transported_to_csv(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "dynamic.xlsx"
            output_dir = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "京东"
            sheet.append(["字段名称", "SKU-DYNAMIC"])
            sheet.append(["商品图", "D:\\materials\\SKU-DYNAMIC.png"])
            sheet.append(["价格3", "新品价"])
            workbook.save(workbook_path)

            clean_data.build_data(
                workbook_path,
                "京东",
                output_dir,
                None,
                profile_id="legacy-v1",
                variant="main-800",
            )
            with (output_dir / "data_全部记录.csv").open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(rows[0]["价格3"], "新品价")

    def test_variant_is_inferred_from_sheet_when_not_explicit(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "tmall.xlsx"
            output_dir = root / "output"
            product = root / "TMALL-750-DEMO.png"
            product.touch()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "现货-750"
            sheet.append([
                "SKU",
                "利益点1",
                "利益点2",
                "预估到手价",
                "价格1",
                "价格2",
                "价格3",
                "规格",
                "卖点",
                "产品",
            ])
            sheet.append([
                "TMALL-750-DEMO",
                "买1减16",
                "买2减36",
                "2件预估均价",
                "71",
                ".9",
                "元/件",
                "茶语卷纸4层200克27卷",
                "200克大胖纸",
                str(product),
            ])
            workbook.save(workbook_path)

            count, errors, _, _, _ = clean_data.build_data(
                workbook_path,
                "现货-750",
                output_dir,
                None,
                profile_id="tmall-positional-v1",
            )
            self.assertEqual((count, errors), (1, 0))
            with (output_dir / "data_全部记录.csv").open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(rows[0]["variant"], "main-750")

    def test_missing_existing_required_field_remains_blocking(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "required-field.xlsx"
            output_dir = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "现货-750"
            sheet.append([
                "商品文件名", "利益点1", "利益点2", "预估到手价", "价格1", "价格2",
                "规格", "卖点", "商品图", "新增文案",
            ])
            sheet.append([
                "TMALL-REQUIRED-DEMO", "买1减16", "买2减36", "2件预估均价", "71", ".9",
                "", "原有卖点", str(product), "新增字段不应改变旧字段校验",
            ])
            workbook.save(workbook_path)

            count, errors, _, _, exception_path = clean_data.build_data(
                workbook_path,
                "现货-750",
                output_dir,
                None,
                profile_id="tmall-positional-v1",
            )
            self.assertEqual((count, errors), (0, 1))
            self.assertIn("字段为空:规格", exception_path.read_text(encoding="utf-8-sig"))

    def test_hygiene_canonical_row_preserves_new_field(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "hygiene.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.CANONICAL_RECORD_ROW_HEADERS + ["价格3"])
            values = {
                "渠道": "天猫官旗", "是否出图": "是", "输出规格": "750",
                "商品文件名": "HYGIENE-V13", "商品图": str(product), "卖点": "透气不闷",
                "片数套": "单套到手", "片数数量": "90片", "到手": "19.9",
                "价格活动价": "29.9", "活动时间": "08/08-08/10", "检查状态": "可出图",
                "价格3": "元/包",
            }
            sheet.append([values.get(header, "") for header in clean_data.CANONICAL_RECORD_ROW_HEADERS] + ["元/包"])
            workbook.save(workbook_path)

            count, errors, data_path, _, _ = clean_data.build_data(
                workbook_path, "出图数据", root / "output", None,
                profile_id="hygiene-tmall-v1.2", variant="main-750",
            )
            self.assertEqual((count, errors), (1, 0))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                record = next(csv.DictReader(handle))
            self.assertEqual(record["商品图"], str(product))
            self.assertEqual(record["价格3"], "元/包")

    def test_hygiene_canonical_row_preserves_new_image_field(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            dynamic_image = root / "dynamic-test.png"
            product.touch()
            dynamic_image.touch()
            workbook_path = root / "hygiene-image.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            extra_headers = ["动态测试字段", "动态测试图片"]
            sheet.append(clean_data.CANONICAL_RECORD_ROW_HEADERS + extra_headers)
            values = {
                "渠道": "天猫官旗", "是否出图": "是", "输出规格": "750",
                "商品文件名": "HYGIENE-V13-IMAGE", "商品图": str(product), "卖点": "透气不闷",
                "片数套": "单套到手", "片数数量": "90片", "到手": "19.9",
                "价格活动价": "29.9", "活动时间": "08/08-08/10", "检查状态": "可出图",
                "动态测试字段": "新增文案", "动态测试图片": str(dynamic_image),
            }
            sheet.append([values.get(header, "") for header in clean_data.CANONICAL_RECORD_ROW_HEADERS + extra_headers])
            workbook.save(workbook_path)

            count, errors, data_path, _, _ = clean_data.build_data(
                workbook_path, "出图数据", root / "output", None,
                profile_id="hygiene-tmall-v1.2", variant="main-750",
            )
            self.assertEqual((count, errors), (1, 0))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                record = next(csv.DictReader(handle))
            self.assertEqual(record["动态测试字段"], "新增文案")
            self.assertEqual(record["动态测试图片"], str(dynamic_image))


if __name__ == "__main__":
    unittest.main()
