#!/usr/bin/env python3
"""Create a single-product workbook for the Windows release E2E test."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--product-image", required=True)
    args = parser.parse_args()

    output_path = Path(args.output).resolve()
    product_image_path = Path(args.product_image).resolve()
    if not product_image_path.is_file():
        raise SystemExit(f"Product image does not exist: {product_image_path}")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "验收数据"
    worksheet.cell(1, 1, "文件名称")
    worksheet.cell(1, 2, "RELEASE-E2E-001")

    fields = [
        ("折扣", "8"),
        ("券名", "限量券"),
        ("满129可用", "满99可用"),
        ("时间", "07.27 00:00 - 07.31 23:59 限时活动，库存有限，售完即止，请以商品页最终信息为准"),
        ("到手", "叠券！3件低至："),
        ("价格1", "8"),
        ("价格2", "折"),
        ("卖点", "悬挂设计抽取方便，厨房清洁随手可取，日常囤货组合装，数量有限先到先得"),
        ("规格", "绒立方厨房抽纸2层180抽3提"),
        ("堆图", str(product_image_path)),
        ("DT17090-24旧", "无.png"),
        ("正式618新旧包装底", "无.png"),
        ("大尺寸", "无.png"),
    ]
    for row_index, (name, value) in enumerate(fields, 2):
        worksheet.cell(row_index, 1, name)
        worksheet.cell(row_index, 2, value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


if __name__ == "__main__":
    main()
