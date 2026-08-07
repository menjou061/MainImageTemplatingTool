"""Regression coverage for the hygiene/Tmall one-row-per-product workbook."""
from __future__ import annotations

import csv
import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
INTERNAL = ROOT / "L0_Windows命令行版" / "_internal"
sys.path.insert(0, str(INTERNAL))
SPEC = importlib.util.spec_from_file_location("l0_clean_data", INTERNAL / "clean_data.py")
assert SPEC and SPEC.loader
clean_data = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(clean_data)


def standard_row(product: Path, output_name: str, **overrides: str) -> list[str]:
    values = {
        "渠道": "天猫官旗",
        "是否出图": "是",
        "输出规格": "750",
        "输出文件名": output_name,
        "商品SKU": output_name,
        "商品系列": "新超薄",
        "模板版式": "小马无侧边",
        "产品图路径": str(product),
        "代言人素材路径": "",
        "主卖点": "100%纯棉柔护",
        "到手套数文案": "2套到手",
        "到手数量": "188片",
        "到手价": "54.1",
        "活动价": "84.5",
        "商品券金额": "",
        "官方立减金额": "20.4",
        "官方立减文案": "官方立减12%",
        "活动时间": "07/07-07/09",
        "赠品版式": "",
        "赠品文案": "",
        "赠品素材路径": "",
        "赠品说明": "",
        "设计备注": "",
        "检查状态": "可出图",
    }
    values.update(overrides)
    return [values.get(header, "") for header in clean_data.STANDARD_RECORD_ROW_HEADERS]


class HygieneRecordRowsTest(unittest.TestCase):
    def test_record_rows_parse_prices_pieces_and_gift_pairs(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            gift_one = root / "gift-one.png"
            gift_two = root / "gift-two.png"
            for path in (product, gift_one, gift_two):
                path.touch()
            workbook_path = root / "hygiene.xlsx"
            output = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "跑批数据"
            sheet.append(clean_data.RECORD_ROW_HEADERS)
            sheet.append(
                [
                    "618预售品", "新超薄", "100%纯棉* 更干爽不黏腻", "*指棉面层",
                    "售卖07/18 00:00:00 - 07/19 23:59:59", str(product), "2套含赠到手242片", "2",
                    "58.3(第2套到手预估)=79.5（活动价）-2（入会领商品券）-19.2（官方立减12%）",
                    "\n".join((str(gift_one), str(gift_two))),
                    "会员0元试用\n新会员0.01元拍下得", "",
                ]
            )
            workbook.save(workbook_path)

            count, errors, data_path, _, error_path = clean_data.build_data(
                workbook_path, "跑批数据", output, None, profile_id="hygiene-tmall-v1.2", variant="main-750"
            )

            self.assertEqual((count, errors), (1, 0))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                record = next(csv.DictReader(handle))
            self.assertEqual(record["卖点"], "100%纯棉* 更干爽不黏腻")
            self.assertEqual(record["备注"], "*指棉面层")
            self.assertEqual(record["片数套"], "2套到手")
            self.assertEqual(record["片数数量"], "242片")
            self.assertEqual(record["到手"], "¥58.3")
            self.assertEqual(record["价格活动价"], "79.5")
            self.assertEqual(record["价格优惠券"], "2")
            self.assertEqual(record["价格立减"], "19.2")
            self.assertEqual(record["赠品图1"], str(gift_one))
            self.assertEqual(record["赠品图2"], str(gift_two))
            self.assertEqual(record["赠品图3"], "")
            self.assertEqual(record["版式组"], "三栏会员")
            self.assertFalse(error_path.read_text(encoding="utf-8-sig").splitlines()[1:])

    def test_price_and_piece_parsers_allow_real_optional_forms(self) -> None:
        without_coupon = clean_data.split_price_bar(
            "41.3(第2套到手预估)=41.3（活动价）-10（官方立减12%）"
        )
        self.assertEqual(without_coupon["到手"], "¥41.3")
        self.assertEqual(without_coupon["价格优惠券"], "")
        self.assertEqual(without_coupon["价格立减"], "10")
        self.assertEqual(clean_data.split_piece_count("2套到手24条"), {"片数套": "2套到手", "片数数量": "24条"})
        self.assertEqual(clean_data.split_piece_count("单套到手88片"), {"片数套": "单套到手", "片数数量": "88片"})
        self.assertEqual(clean_data.split_piece_count("含赠到手88片"), {"片数套": "含赠到手", "片数数量": "88片"})

    def test_standard_sheet_uses_output_name_and_skips_hidden_or_disabled_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "standard.xlsx"
            output = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.STANDARD_RECORD_ROW_HEADERS)
            base = {
                "渠道": "天猫官旗",
                "是否出图": "是",
                "输出规格": "750",
                "输出文件名": "sku-750",
                "商品SKU": "sku",
                "商品系列": "新超薄",
                "模板版式": "小马无侧边",
                "产品图路径": str(product),
                "代言人素材路径": "",
                "主卖点": "100%纯棉柔护",
                "到手套数文案": "2套到手",
                "到手数量": "188片",
                "到手价": "54.1",
                "活动价": "84.5",
                "商品券金额": "",
                "官方立减金额": "20.4",
                "官方立减文案": "官方立减12%",
                "活动时间": "07/07-07/09",
                "赠品版式": "",
                "赠品文案": "",
                "赠品素材路径": "",
                "赠品说明": "",
                "设计备注": "",
                "检查状态": "可出图",
            }
            sheet.append([base.get(header, "") for header in clean_data.STANDARD_RECORD_ROW_HEADERS])
            disabled = dict(base, **{"是否出图": "否", "输出文件名": "skip-disabled"})
            sheet.append([disabled.get(header, "") for header in clean_data.STANDARD_RECORD_ROW_HEADERS])
            hidden = dict(base, **{"输出文件名": "skip-hidden"})
            sheet.append([hidden.get(header, "") for header in clean_data.STANDARD_RECORD_ROW_HEADERS])
            sheet.row_dimensions[4].hidden = True
            workbook.save(workbook_path)

            count, errors, data_path, _, _ = clean_data.build_data(
                workbook_path, "出图数据", output, None, profile_id="hygiene-tmall-v1.2", variant="main-750"
            )

            self.assertEqual((count, errors), (1, 0))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                record = next(csv.DictReader(handle))
            self.assertEqual(record["商品文件名"], "sku-750")
            self.assertEqual(record["版式组"], "三栏会员")
            self.assertEqual(record["价格优惠券"], "")
            self.assertEqual(record["价格立减"], "20.4")

    def test_channel_mismatch_is_recorded_while_unselected_rows_stay_silent(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "channels.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.STANDARD_RECORD_ROW_HEADERS)
            sheet.append(standard_row(product, "wrong-channel", **{"渠道": "京东自营"}))
            sheet.append(standard_row(product, "disabled", **{"渠道": "京东自营", "是否出图": "否"}))
            sheet.append(standard_row(product, "other-spec", **{"输出规格": "800"}))
            sheet.append(standard_row(product, "valid"))
            workbook.save(workbook_path)

            count, errors, data_path, all_data_path, error_path = clean_data.build_data(
                workbook_path,
                "出图数据",
                root / "output",
                None,
                profile_id="hygiene-tmall-v1.2",
                variant="main-750",
            )

            self.assertEqual((count, errors), (1, 2))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual([row["商品文件名"] for row in csv.DictReader(handle)], ["valid"])
            with all_data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(
                    [row["商品文件名"] for row in csv.DictReader(handle)],
                    ["wrong-channel", "other-spec", "valid"],
                )
            with error_path.open(encoding="utf-8-sig", newline="") as handle:
                errors_by_product = {row["商品文件名"]: row for row in csv.DictReader(handle)}
            self.assertEqual(set(errors_by_product), {"wrong-channel", "other-spec"})
            self.assertEqual(errors_by_product["wrong-channel"]["错误码"], "E_CHANNEL_MISMATCH")
            self.assertIn("京东自营", errors_by_product["wrong-channel"]["异常详情"])
            self.assertIn("当前所选渠道", errors_by_product["wrong-channel"]["建议动作"])
            self.assertEqual(errors_by_product["other-spec"]["错误码"], "E_OUTPUT_SPEC_MISMATCH")

    def test_standard_contract_rejects_unknown_status_and_gift_layout(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "contract.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.STANDARD_RECORD_ROW_HEADERS)
            sheet.append(standard_row(product, "bad-status", **{"检查状态": "待确认"}))
            sheet.append(standard_row(product, "bad-gift-layout", **{"赠品版式": "顶部赠品"}))
            sheet.append(standard_row(product, "missing-spec", **{"输出规格": ""}))
            sheet.append(
                standard_row(
                    product,
                    "combined-errors",
                    **{"输出规格": "", "检查状态": "待确认", "赠品版式": "顶部赠品"},
                )
            )
            workbook.save(workbook_path)

            count, errors, data_path, all_data_path, error_path = clean_data.build_data(
                workbook_path,
                "出图数据",
                root / "output",
                None,
                profile_id="hygiene-tmall-v1.2",
                variant="main-750",
            )

            self.assertEqual((count, errors), (0, 4))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(list(csv.DictReader(handle)), [])
            with all_data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(
                    [row["商品文件名"] for row in csv.DictReader(handle)],
                    ["bad-status", "bad-gift-layout", "missing-spec", "combined-errors"],
                )
            with error_path.open(encoding="utf-8-sig", newline="") as handle:
                errors_by_product = {row["商品文件名"]: row for row in csv.DictReader(handle)}
            self.assertEqual(errors_by_product["bad-status"]["错误码"], "E_CHECK_STATUS_INVALID")
            self.assertEqual(errors_by_product["bad-gift-layout"]["错误码"], "E_GIFT_LAYOUT_UNSUPPORTED")
            self.assertEqual(errors_by_product["missing-spec"]["错误码"], "E_OUTPUT_SPEC_MISSING")
            combined_detail = errors_by_product["combined-errors"]["异常详情"]
            self.assertIn("E_OUTPUT_SPEC_MISSING", combined_detail)
            self.assertIn("E_CHECK_STATUS_INVALID", combined_detail)
            self.assertIn("E_GIFT_LAYOUT_UNSUPPORTED", combined_detail)

    def test_limit_keeps_interleaved_exceptions_within_selected_records(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            product.touch()
            workbook_path = root / "limited.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.STANDARD_RECORD_ROW_HEADERS)
            sheet.append(standard_row(product, "invalid-first", **{"赠品文案": "赠品一"}))
            sheet.append(standard_row(product, "valid-second"))
            sheet.append(standard_row(product, "invalid-after-limit", **{"渠道": "京东自营"}))
            workbook.save(workbook_path)

            count, errors, data_path, all_data_path, error_path = clean_data.build_data(
                workbook_path,
                "出图数据",
                root / "output",
                None,
                limit=2,
                profile_id="hygiene-tmall-v1.2",
                variant="main-750",
            )

            self.assertEqual((count, errors), (1, 1))
            with data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual([row["商品文件名"] for row in csv.DictReader(handle)], ["valid-second"])
            with all_data_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(
                    [row["商品文件名"] for row in csv.DictReader(handle)],
                    ["invalid-first", "valid-second"],
                )
            with error_path.open(encoding="utf-8-sig", newline="") as handle:
                error_rows = list(csv.DictReader(handle))
            self.assertEqual([row["商品文件名"] for row in error_rows], ["invalid-first"])
            self.assertEqual(error_rows[0]["错误码"], "E_GIFT_PAIR_MISMATCH")

    def test_gift_pair_mismatch_is_recorded_with_actionable_code(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            gift = root / "gift.png"
            product.touch()
            gift.touch()
            workbook_path = root / "mismatch.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.RECORD_ROW_HEADERS)
            sheet.append([
                "618", "新超薄", "卖点", "", "07/07-07/09", str(product), "2套到手88片", "2",
                "41.3(到手)=51.3（活动价）-10（官方立减）", str(gift), "赠品一\n赠品二", "",
            ])
            workbook.save(workbook_path)
            count, errors, _, _, error_path = clean_data.build_data(
                workbook_path, "出图数据", root / "output", None,
                profile_id="hygiene-tmall-v1.2", variant="main-750",
            )
            self.assertEqual((count, errors), (0, 1))
            with error_path.open(encoding="utf-8-sig", newline="") as handle:
                error = next(csv.DictReader(handle))
            self.assertEqual(error["错误码"], "E_GIFT_PAIR_MISMATCH")
            self.assertIn("一一对应", error["建议动作"])

    def test_750_accepts_three_gift_cards_before_photoshop_runs(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            product = root / "product.png"
            gifts = [root / f"gift-{index}.png" for index in range(1, 4)]
            for path in [product] + gifts:
                path.touch()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(clean_data.RECORD_ROW_HEADERS)
            sheet.append([
                "618", "新超薄", "卖点", "", "07/07-07/09", str(product), "2套到手88片", "2",
                "41.3(到手)=51.3（活动价）-10（官方立减）",
                "\n".join(str(path) for path in gifts), "赠品一\n赠品二\n赠品三", "",
            ])
            workbook_path = root / "three-gifts.xlsx"
            workbook.save(workbook_path)

            count, errors, _, _, error_path = clean_data.build_data(
                workbook_path, "出图数据", root / "output", None,
                profile_id="hygiene-tmall-v1.2", variant="main-750",
            )

            self.assertEqual((count, errors), (1, 0))
            with error_path.open(encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(list(csv.DictReader(handle)), [])


if __name__ == "__main__":
    unittest.main()
