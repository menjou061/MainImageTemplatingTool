"""Regression coverage for product listing in designer-facing workbooks."""
from __future__ import annotations

import importlib.util
import io
import re
import tempfile
import unittest
import zipfile
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "L0_Windows命令行版" / "_internal" / "l0_list_sheets.py"
SPEC = importlib.util.spec_from_file_location("l0_list_sheets", SCRIPT)
assert SPEC and SPEC.loader
sheet_listing = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(sheet_listing)


class SheetListingTest(unittest.TestCase):
    def test_standard_hygiene_sheet_lists_only_enabled_visible_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "standard.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(["是否出图", "输出规格", "输出文件名", "产品图路径", "主卖点"])
            sheet.append(["是", "750", "sku-visible", r"C:\materials\one.png", "卖点"])
            sheet.append(["否", "750", "sku-disabled", r"C:\materials\two.png", "卖点"])
            sheet.append(["是", "750", "sku-hidden", r"C:\materials\three.png", "卖点"])
            sheet.row_dimensions[4].hidden = True
            workbook.save(workbook_path)

            output = io.StringIO()
            with redirect_stdout(output):
                result = sheet_listing.list_products(workbook_path, "出图数据")

            self.assertEqual(result, 0)
            self.assertEqual(output.getvalue().splitlines(), ["sku-visible"])

    def test_standard_hygiene_sheet_filters_current_channel_and_variant(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "mixed.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(["渠道", "是否出图", "输出规格", "输出文件名", "产品图路径", "主卖点"])
            sheet.append(["天猫官旗", "是", "750", "sku-current", r"C:\materials\one.png", "卖点"])
            sheet.append(["天猫官旗", "是", "800", "sku-other-size", r"C:\materials\two.png", "卖点"])
            sheet.append(["京东自营", "是", "750", "sku-other-channel", r"C:\materials\three.png", "卖点"])
            workbook.save(workbook_path)

            output = io.StringIO()
            with redirect_stdout(output):
                result = sheet_listing.list_products(
                    workbook_path,
                    "出图数据",
                    profile_id="hygiene-tmall-v1.2",
                    variant="main-750",
                )

            self.assertEqual(result, 0)
            self.assertEqual(output.getvalue().splitlines(), ["sku-current"])

    def test_streamed_sheet_without_dimension_metadata_does_not_crash(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            fixture = root / "dimensionless.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "出图数据"
            sheet.append(["是否出图", "输出规格", "输出文件名", "产品图路径", "主卖点"])
            sheet.append(["是", "750", "sku-dimensionless", r"C:\materials\one.png", "卖点"])
            workbook.save(source)
            with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(fixture, "w") as output_archive:
                for item in archive.infolist():
                    payload = archive.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml":
                        payload = re.sub(br'<dimension ref="[^"]+"\s*/>', b"", payload, count=1)
                    output_archive.writestr(item, payload)

            output = io.StringIO()
            with redirect_stdout(output):
                result = sheet_listing.list_products(fixture, "出图数据")
            self.assertEqual(result, 0)
            self.assertEqual(output.getvalue().splitlines(), ["sku-dimensionless"])


if __name__ == "__main__":
    unittest.main()
