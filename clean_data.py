#!/usr/bin/env python3
"""Convert a horizontal e-commerce Excel matrix into UTF-8 CSV rows.

No mapping file is used. The first-column variable names in the selected
worksheet become CSV columns; the small set of standard name normalisations
below mirrors the approved PSD naming convention.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


NAME_MAP = {
    "时间": "活动时间",
    "满129可用": "券门槛",
    "堆图": "商品图",
    "大尺寸": "大尺寸图",
    "DT17090-24旧": "旧包装图",
    "正式618新旧包装底": "新旧包装底图",
}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".psd"}
IMAGE_FIELDS = {"商品图", "大尺寸图", "旧包装图", "新旧包装底图"}
SWITCH_IMAGE_MAP = {
    "展示大尺寸": "大尺寸图",
    "展示新旧包装": ("旧包装图", "新旧包装底图"),
}
STANDARD_COLUMNS = [
    "商品文件名",
    "商品图",
    "折扣",
    "券名",
    "券门槛",
    "优惠券开关",
    "活动时间",
    "到手",
    "价格1",
    "价格2",
    "卖点",
    "规格",
    "展示大尺寸",
    "大尺寸图",
    "展示新旧包装",
    "旧包装图",
    "新旧包装底图",
    "预检异常",
]


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def has_newline(value: Any) -> bool:
    return isinstance(value, str) and ("\n" in value or "\r" in value)


def is_display_image_formula(value: Any) -> bool:
    return isinstance(value, str) and value.strip().lower().startswith("=_xlfn.dispimg")


def field_name(value: Any) -> str:
    raw = as_text(value)
    return NAME_MAP.get(raw, raw)


def local_image_name(value: str) -> str:
    """Discard network/absolute paths so JSX always searches local material."""
    candidate = value.replace("\\", "/")
    suffix = Path(candidate).suffix.lower()
    if suffix in IMAGE_SUFFIXES:
        return Path(candidate).name
    return value


def split_price(value: Any) -> tuple[str, str]:
    text = as_text(value).replace("￥", "").replace(" ", "")
    match = re.fullmatch(r"(\d+)(\.\d+)?", text)
    if not match:
        return text, ""
    return match.group(1), match.group(2) or ""


def price_format_error(record: dict[str, str]) -> str:
    """Only require both PSD price layers to have a value.

    ``价格2`` is presentation copy, not a numeric field.  It can be a decimal
    suffix, unit, discount text, or any other text that should follow 价格1.
    """
    price1 = as_text(record.get("价格1", ""))
    price2 = as_text(record.get("价格2", ""))
    if not price1 or not price2:
        return "价格1、价格2必须完整填写"
    return ""


def is_disabled_image(value: Any) -> bool:
    text = as_text(value)
    return is_blank(text) or local_image_name(text).lower() in {"无", "无.png", "none", "null"}


def normalize_price_part(value: Any, part: str) -> str:
    text = as_text(value)
    if part == "价格2" and re.fullmatch(r"0\.\d+", text):
        return text[1:]
    return text


def validate_variable_name(name: str) -> None:
    if name and (name[0].isdigit() or re.search(r"[\\/:,;\r\n]", name)):
        raise ValueError(f"变量名不符合 Photoshop/CSV 规范：{name}")


def choose_inputs(args: argparse.Namespace) -> tuple[Path, str, Path]:
    workbook_path = Path(args.xlsx).expanduser() if args.xlsx else None
    if workbook_path is None:
        try:
            from tkinter import Tk, filedialog

            root = Tk()
            root.withdraw()
            picked = filedialog.askopenfilename(title="选择 Excel 变量表", filetypes=[("Excel", "*.xlsx")])
            root.destroy()
            workbook_path = Path(picked) if picked else None
        except Exception:
            workbook_path = Path(input("请输入 Excel 文件路径：").strip())
    if not workbook_path or not workbook_path.is_file():
        raise SystemExit("未找到 Excel 文件。")

    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        raise SystemExit("工作簿中没有可用的可见 Sheet。")
    sheet = args.sheet
    if not sheet:
        print("可用 Sheet：")
        for index, title in enumerate(visible_sheets, 1):
            print(f"  {index}. {title}")
        selection = input("请输入 Sheet 编号或名称：").strip()
        sheet = visible_sheets[int(selection) - 1] if selection.isdigit() else selection
    if sheet not in visible_sheets:
        raise SystemExit("所选 Sheet 不存在、已隐藏，或是 WPS 内嵌图片索引 Sheet。")

    output_dir = Path(args.output_dir).expanduser() if args.output_dir else workbook_path.parent / f"套版数据_{sheet}"
    return workbook_path, sheet, output_dir


def source_variables(ws) -> list[tuple[int, str]]:
    variables: list[tuple[int, str]] = []
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row, 1).value
        if is_blank(raw) or is_display_image_formula(raw):
            continue
        raw_name = as_text(raw)
        # These are worksheet notes, not data fields for a product row.
        if raw_name.startswith("\\\\") or raw_name.startswith("//"):
            continue
        name = field_name(raw_name)
        if name:
            validate_variable_name(name)
            variables.append((row, name))
    return variables


def row_from_column(ws, column: int, variables: list[tuple[int, str]]) -> dict[str, str]:
    row: dict[str, str] = {}
    for source_row, target_name in variables:
        value = ws.cell(source_row, column).value
        if has_newline(value):
            row["预检异常"] = append_issue(row.get("预检异常", ""), "脏数据")
        text = normalize_price_part(value, target_name)
        row[target_name] = local_image_name(text)

    if "价格" in row and is_blank(row.get("价格1")) and is_blank(row.get("价格2")):
        row["价格1"], row["价格2"] = split_price(row["价格"])
    elif is_blank(row.get("价格2")) and re.fullmatch(r"\d+\.\d+", row.get("价格1", "")):
        row["价格1"], row["价格2"] = split_price(row["价格1"])

    coupon_fields = [row.get("折扣", ""), row.get("券名", ""), row.get("券门槛", "")]
    filled = [not is_blank(value) for value in coupon_fields]
    if not any(filled):
        row["优惠券开关"] = "否"
    else:
        row["优惠券开关"] = "是"
        if not all(filled):
            row["预检异常"] = append_issue(row.get("预检异常", ""), "字段为空")

    for switch_name, image_field in SWITCH_IMAGE_MAP.items():
        if isinstance(image_field, tuple):
            visible = all(not is_disabled_image(row.get(field, "")) for field in image_field)
        else:
            visible = not is_disabled_image(row.get(image_field, ""))
        row[switch_name] = "是" if visible else "否"
    return row


def append_issue(existing: str, issue: str) -> str:
    return issue if not existing else existing + ";" + issue


def quality_score(record: dict[str, str]) -> int:
    return sum(1 for key, value in record.items() if key not in {"预检异常", "优惠券开关"} and not is_blank(value))


def material_index(folder: Path | None) -> set[str]:
    if folder is None:
        return set()
    return {path.name.lower() for path in folder.rglob("*") if path.is_file()}


def add_material_precheck(record: dict[str, str], available_files: set[str]) -> None:
    if not available_files:
        return
    for field in IMAGE_FIELDS:
        image_name = record.get(field, "")
        if not is_disabled_image(image_name) and image_name.lower() not in available_files:
            record["预检异常"] = append_issue(record.get("预检异常", ""), f"缺图:{field}")


def add_filename_precheck(record: dict[str, str]) -> None:
    for field in IMAGE_FIELDS:
        image_name = local_image_name(record.get(field, ""))
        if is_disabled_image(image_name):
            continue
        if "m²" in image_name or "²" in image_name:
            record["预检异常"] = append_issue(record.get("预检异常", ""), f"文件名特殊字符:{field}")


def exception_record(record: dict[str, str], product: str, column: int, issue: str, detail: str) -> dict[str, str]:
    result = dict(record)
    result.update({"商品文件名": product, "源列": str(column), "异常类型": issue, "异常详情": detail})
    return result


def write_csv(path: Path, fieldnames: list[str], records: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def build_data(
    workbook_path: Path,
    sheet_name: str,
    output_dir: Path,
    product_filter: str | None,
    assets_dir: str | None,
    limit: int | None = None,
) -> tuple[int, int, Path, Path]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    ws = workbook[sheet_name]
    if ws.sheet_state != "visible" or ws.title == "WpsReserved_CellImgList":
        raise ValueError("不能处理隐藏 Sheet 或 WPS 内嵌图片索引 Sheet。")
    local_assets = Path(assets_dir).expanduser() if assets_dir else None
    if local_assets and not local_assets.is_dir():
        raise ValueError("指定的本地素材文件夹不存在。")
    available_files = material_index(local_assets)

    variables = source_variables(ws)
    candidates: list[tuple[int, str, dict[str, str]]] = []
    exceptions: list[dict[str, str]] = []
    for column in range(2, ws.max_column + 1):
        header = ws.cell(1, column).value
        if is_blank(header) or is_display_image_formula(header):
            continue
        product = as_text(header)
        if has_newline(header):
            record = row_from_column(ws, column, variables)
            exceptions.append(exception_record(record, product, column, "脏数据", "商品文件名包含换行符"))
            continue
        if product_filter and product != product_filter:
            continue
        record = row_from_column(ws, column, variables)
        add_filename_precheck(record)
        add_material_precheck(record, available_files)
        if "脏数据" in record.get("预检异常", ""):
            exceptions.append(exception_record(record, product, column, "脏数据", "任一单元格包含换行符"))
            continue
        price_error = price_format_error(record)
        if price_error:
            record["预检异常"] = append_issue(record.get("预检异常", ""), "价格格式异常")
            exceptions.append(exception_record(record, product, column, "价格格式异常", price_error))
            continue
        candidates.append((column, product, record))

    by_product: dict[str, list[tuple[int, str, dict[str, str]]]] = defaultdict(list)
    for candidate in candidates:
        by_product[candidate[1]].append(candidate)

    clean_records: list[dict[str, str]] = []
    clean_record_sources: dict[str, int] = {}
    for product, group in by_product.items():
        group.sort(key=lambda item: (-quality_score(item[2]), item[0]))
        kept = group[0]
        kept_record = dict(kept[2])
        kept_record["商品文件名"] = product
        clean_records.append(kept_record)
        clean_record_sources[product] = kept[0]
        precheck = kept_record.get("预检异常", "")
        for marker, issue_type in (
            ("缺图:", "缺图"),
            ("字段为空", "字段为空"),
            ("文件名特殊字符:", "文件名特殊字符"),
        ):
            if marker in precheck:
                exceptions.append(exception_record(kept_record, product, kept[0], issue_type, precheck))
        for duplicate in group[1:]:
            exceptions.append(
                exception_record(
                    duplicate[2],
                    product,
                    duplicate[0],
                    "重复商品名",
                    f"保留了信息更完整的第 {kept[0]} 列；当前为第 {duplicate[0]} 列。",
                )
            )

    if limit is not None:
        if limit < 1:
            raise ValueError("--limit 必须是大于 0 的整数。")
        clean_records = clean_records[:limit]
        selected_products = {record["商品文件名"] for record in clean_records}
        maximum_source_column = max(clean_record_sources[product] for product in selected_products)
        exceptions = [
            record for record in exceptions
            if int(record.get("源列", 0) or 0) <= maximum_source_column
        ]
    seen_columns = set(STANDARD_COLUMNS)
    discovered_columns = []
    for record in clean_records + exceptions:
        for key in record:
            if key not in seen_columns and key not in {"源列", "异常类型", "异常详情"}:
                seen_columns.add(key)
                discovered_columns.append(key)
    data_columns = STANDARD_COLUMNS + discovered_columns
    for record in clean_records:
        for column in data_columns:
            record.setdefault(column, "")

    output_dir.mkdir(parents=True, exist_ok=True)
    data_path = output_dir / "data.csv"
    error_path = output_dir / "异常记录.csv"
    write_csv(data_path, data_columns, clean_records)
    error_columns = ["商品文件名", "源列", "异常类型", "异常详情"] + [column for column in data_columns if column != "商品文件名"]
    write_csv(error_path, error_columns, exceptions)
    return len(clean_records), len(exceptions), data_path, error_path


def main() -> None:
    parser = argparse.ArgumentParser(description="将横向商品变量表清洗为套版 data.csv（UTF-8）。")
    parser.add_argument("--xlsx", help="Excel 变量表路径；省略时弹窗选择。")
    parser.add_argument("--sheet", help="要处理的可见 Sheet 名称；省略时交互选择。")
    parser.add_argument("--output-dir", help="data.csv 与异常记录.csv 输出文件夹。")
    parser.add_argument("--assets-dir", help="本地素材文件夹，仅验证其存在；CSV 内始终保存素材文件名。")
    parser.add_argument("--product", help="仅导出指定商品文件名，用于单条试跑。")
    parser.add_argument("--limit", type=int, help="仅导出前 N 个去重后的商品，用于分批试跑。")
    args = parser.parse_args()
    workbook_path, sheet_name, output_dir = choose_inputs(args)
    try:
        count, errors, data_path, error_path = build_data(
            workbook_path, sheet_name, output_dir, args.product, args.assets_dir, args.limit
        )
    except (OSError, ValueError) as error:
        raise SystemExit(f"处理失败：{error}")
    print(f"完成：有效记录 {count} 条，异常 {errors} 条。")
    print(f"数据文件：{data_path}")
    print(f"异常记录：{error_path}")


if __name__ == "__main__":
    main()
